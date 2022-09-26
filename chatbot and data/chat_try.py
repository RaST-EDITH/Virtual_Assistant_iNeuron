import openpyxl as oxl

file = open(r"C:\Users\ASUS\bot.txt","w", encoding="utf-8")

wb=oxl.load_workbook(r"C:\Users\ASUS\raw.xlsx")
ws=wb.active
temp=[]
size = ws.max_row+1

for i in range(4,size):

    x=ws[f"S{i}"].value
    z=x.split("\n")
    
    while ''  in z :
        z.remove('')
    
    if len(z) > 2 :
        z=z[1:]

        while ''  in z :
            z.remove('')

        # print(z)
        for a in z :
            ch="How may I assist you"
            nouse = [ "hello" , "hi" , "hi sir" , "hello sir" , ""]
            q=a.find(":")
            line = a[q+1:]
            if a[q+1] == " " :
                line = a[q+2:]
            if a[q-5:q] == "https" :
                line = a[q-5:]
            if ch.lower() in a.lower() :
                continue
            if a == "Hi" :
                continue
            if a.casefold() in nouse :
                continue

            file.write(line+"\n")
            # file.write("\n")
        file.write("\n")

file.close()
